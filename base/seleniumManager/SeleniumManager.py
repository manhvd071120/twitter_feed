
import undetected_chromedriver.v2 as uc
from time import sleep
import openpyxl
from selenium.common.exceptions import NoSuchElementException
from selenium.webdriver import ActionChains, Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.wait import WebDriverWait
from selenium.common.exceptions import TimeoutException
from modules.seleniumIDE.autoSeleModel import ListTwitter
import random
from base.model import ChromeProfile
from base.utils import Utils
from win32api import GetFileVersionInfo, HIWORD
import logging
import os
TimeOut = 15

logging.basicConfig(filename=f'./error.log', level=logging.INFO, format='%(asctime)s - %(message)s')
def createNewChrome(chromeProfile: ChromeProfile, proxy, numThreads, posThread, configInfo):
    options = uc.ChromeOptions()
    # options.binary_location = "F:\\brave-portable\\brave-portable.exe"
    if not os.path.exists(chromeProfile.profilePath):
        os.makedirs(chromeProfile.profilePath)
    if chromeProfile.chromePath != "":
        options.binary_location = f'{chromeProfile.chromePath}'
    options.user_data_dir = chromeProfile.profilePath
    # End

    chrome_version = HIWORD(GetFileVersionInfo(options.binary_location, "\\").get('FileVersionMS'))
    if chrome_version == 92:
        chrome_driver = './chromedriver92.exe'
    elif chrome_version == 85:
        chrome_driver = './chromedriver85.exe'
    elif chrome_version == 95:
        chrome_driver = './chromedriver95.exe'
    elif chrome_version == 96:
        chrome_driver = './chromedriver96.exe'
    else:
        chrome_driver = './chromedriver93.exe'

    # End
    chrome_locale = 'en-us'
    options.add_argument("--lang={}".format(chrome_locale))
    # Dis
    options.add_argument('--disable-session-crashed-bubble')
    options.add_argument('--disable-application-cache')
    # Pop-up
    options.add_argument('--disable-popup-blocking')
    # End
    # Extension
    if configInfo['modeNetwork'] == 'Proxy':
        if chromeProfile.ipaddress != "":
            options.add_argument(f'--proxy-server={chromeProfile.ipaddress}')
    else:
        options.add_argument(f'--proxy-server={proxy}')

    # if chromeProfile.webRtcPath != "":
    exPath = getListExtension(chromeProfile.webRtcPath)
    extension_path = f'{exPath}'
    #     options.add_argument('--disable-extensions')
    #     options.add_argument(f'--disable-extensions-except={extension_path}')
    options.add_argument(f'--load-extension={extension_path}')
    # End

    # User Agent
    if chromeProfile.userAgent != "":
        options.add_argument(f'user-agent=={chromeProfile.userAgent}')
    # End
    options.add_argument('--no-first-run --no-service-autorun --password-store=basic')
    driver = uc.Chrome(executable_path=chrome_driver, options=options)
    # try:
    #     width, height = pyautogui.size()
    #     width_window = width / numThreads
    #     index_window = posThread * width_window
    #     driver.set_window_rect(index_window, 0, width_window, 1000)
    # except:
    # driver.set_window_rect(100, 0, 400, 800)
    return driver
# Mat khau

ListTwitters = []

def check_susspend(chromeDriver,configInfo, profileInfo, listLock):
    try:
        sleep(random.randint(3, 10))
        check = random.randint(1, 3)
        if check == 2:
            for i in range(random.randint(5, 12)):
                for j in range(random.randint(40, 60)):
                    scrollBy(chromeDriver, 0, random.randint(10, 20))
                sleep(random.randint(5, 20))
        for i in range(5):
            btnProfile = findElementByXpath(chromeDriver, '//a[@aria-label="Profile"]', 5)
            if btnProfile != "":
                moveToElementandClick(chromeDriver, btnProfile)
                sleep(random.randint(5, 10))
                break
            elif i == 4:
                return
            else:
                btnProfile = findElementByXpath(chromeDriver, '//a[@aria-label="Hồ sơ"]', 5)
                moveToElementandClick(chromeDriver, btnProfile)
                sleep(random.randint(5, 10))
                break
        check_following = findElementsByXpath(chromeDriver,'//span[@class="css-901oao css-16my406 r-18jsvk2 r-poiln3 r-b88u0q r-bcqeeo r-qvutc0"]')
        following = str(check_following[0].text)
        for i in range(random.randint(0, 3)):
            for j in range(random.randint(40, 60)):
                scrollBy(chromeDriver, 0, random.randint(10, 20))
            sleep(random.randint(5, 20))
        if int(following) == 0:
            ListTwitters.append(ListTwitter(profileInfo.get("Email"), 'susspend', None))
            exit()
        sleep(random.randint(1, 10))
        for i in range(5):
            btnHome = findElementByXpath(chromeDriver, '//a[@href="/home"]', 5)
            if btnHome != "":
                clickButton(btnHome)
                sleep(random.randint(1, 10))
                break
            elif i == 4:
                goToWebsite(chromeDriver, 'https://twitter.com/home')
                sleep(random.randint(1, 10))
            else:
                btnHome = findElementByXpath(chromeDriver, '//a[@aria-label="Trang chủ"]', 5)
                if btnHome != "":
                    clickButton(btnHome)
                    sleep(random.randint(1, 10))
                    break
    except:
        pass

def twitter_feed(chromeDriver,configInfo, profileInfo, listLock):
    btnContinue1 = findElementByXpath(chromeDriver, '//span[text()="Sign in"]', random.randint(3, 5))
    if btnContinue1 != "":
        ListTwitters.append(ListTwitter(profileInfo.get("Email"), 'login', None))
        return
    else:
        btnContinue2 = findElementByXpath(chromeDriver, '//span[text()="Đăng nhập"]', random.randint(3, 5))
        if btnContinue2 != "":
            ListTwitters.append(ListTwitter(profileInfo.get("Email"), 'login', None))
            return
    btnCheckpoint = findElementByXpath(chromeDriver, '//input[@value="Start"]', random.randint(3, 5))
    if btnCheckpoint != "":
        ListTwitters.append(ListTwitter(profileInfo.get("Email"), 'check point', None))
        return
    else:
        btnCheckpoint = findElementByXpath(chromeDriver, '//input[@value="Bắt đầu"]', random.randint(3, 5))
        if btnCheckpoint != "":
            ListTwitters.append(ListTwitter(profileInfo.get("Email"), 'check point', None))
            return
    btnCheckpoint1 = findElementByXpath(chromeDriver, '//input[@value="Start"]', random.randint(3, 5))
    if btnCheckpoint1 != "":
        ListTwitters.append(ListTwitter(profileInfo.get("Email"), 'check point', None))
        return
    else:
        btnCheckpoint1 = findElementByXpath(chromeDriver, '//input[@value="Bắt đầu"]', random.randint(3, 5))
        if btnCheckpoint1 != "":
            ListTwitters.append(ListTwitter(profileInfo.get("Email"), 'check point', None))
            return
    check_susspend(chromeDriver,configInfo, profileInfo, listLock)
    sleep(random.randint(1,10))
    for i in range(random.randint(5, 12)):
        try:
            twitter_scroll_Down(chromeDriver)
            count = random.randint(1, 5)
            clickTweet(chromeDriver)
            if count == 2:
                twitter_scroll_Up(chromeDriver)
        except:
            pass
    try:
        time = int(profileInfo.get('Time Run')) + 1
        ListTwitters.append(ListTwitter(profileInfo.get("Email"), 'running', time))
    except:
        ListTwitters.append(ListTwitter(profileInfo.get("Email"), 'running', 1))

def twitter_scroll_Down(chromeDriver):
    try:
        for i in range(random.randint(6, 16)):
            for j in range(random.randint(40, 60)):
                scrollBy(chromeDriver, 0, random.randint(10, 20))
            sleep(random.randint(5, 35))
    except:
        pass

def clickTweet(chromeDriver):
    try:
        listPost = findElementsByXpath(chromeDriver,'//div[@class="css-901oao r-18jsvk2 r-37j5jr r-a023e6 r-16dba41 r-rjixqe r-bcqeeo r-bnwqim r-qvutc0"]',5)
        if listPost != "":
            current_window = chromeDriver.current_window_handle
            moveToElementandClick(chromeDriver, listPost[random.randint(0, len(listPost) - 1)])
            try:
                WebDriverWait(chromeDriver, 5).until(EC.number_of_windows_to_be(2))
                sleep(random.randint(2, 3))
                for chrome_window in chromeDriver.window_handles:
                    chromeDriver.switch_to.window(chrome_window)
                    sleep(random.randint(2, 3))
                    if chrome_window != current_window:
                        chromeDriver.close()
                        sleep(random.randint(2, 3))
                        chromeDriver.switch_to.window(current_window)
            except TimeoutException:
                pass
            for i in range(random.randint(0, 5)):
                if i == 0:
                    sleep(random.randint(1, 10))
                for j in range(random.randint(40, 55)):
                    scrollBy(chromeDriver, 0, random.randint(10, 20))
                sleep(random.randint(5, 20))
            sleep(random.randint(3, 5))
            for i in range(5):
                btnHome = findElementByXpath(chromeDriver, '//a[@href="/home"]', 5)
                if btnHome != "":
                    clickButton(btnHome)
                    break
                elif i == 4:
                    goToWebsite(chromeDriver, 'https://twitter.com/home')
                    sleep(random.randint(1, 10))
                else:
                    btnHome = findElementByXpath(chromeDriver, '//a[@aria-label="Trang chủ"]', 5)
                    if btnHome != "":
                        clickButton(btnHome)
                        sleep(random.randint(3, 5))
                        break
    except:
        pass

def twitter_scroll_Up(chromeDriver):
    try:
        for i in range(random.randint(5, 12)):
            for j in range(random.randint(30, 45)):
                scrollBy(chromeDriver, 0, random.randint(-20, -10))
            sleep(random.randint(5, 20))
    except:
        pass

def saveListTwtter(configInfo, listTwitter):
    try:
        wb = openpyxl.load_workbook(configInfo['proxyPath'])
        ws = wb.active
        row = ws.max_row
        column_list = [cell.value for cell in ws[1]]
        count = 0
        for i in range(1, row + 1):
            if i == 1:
                for j in range(len(column_list)):
                    if column_list[j] == 'Email':
                        inEmail = j + 1
                    if column_list[j] == "Status":
                        inStatus = j + 1
                    if column_list[j] == "Time Run":
                        inTimeRun = j + 1
            else:
                if count == len(listTwitter):
                    break
                for items in listTwitter:
                    if ws.cell(i, inEmail).value == items.Email:
                        ws.cell(i,inTimeRun).value = items.TimeRun
                        ws.cell(i, inStatus).value = items.Status
                        count += 1
        wb.save(filename=configInfo['proxyPath'])
        wb.close()
    except Exception:
        print('file ghi loi')

def ResetFileExcel(filePath):
    try:
        wb = openpyxl.load_workbook(filePath)
        ws = wb.active
        row = ws.max_row
        column_list = [cell.value for cell in ws[1]]

        for i in range(1, row + 1):
            if i == 1:
                for j in range(len(column_list)):
                    if column_list[j] == 'Status':
                        inStatus = j + 1
                    if column_list[j] == 'Time Run':
                        inTimeRun = j + 1
            elif i >= 2:
                ws.cell(i, inStatus).value = ''
                ws.cell(i, inTimeRun).value = 0
        wb.save(filename=filePath)
        wb.close()
    except Exception:
        print('Reset file ghi loi')

def loginProxy(chromeDriver, infoProxy):
    waitLoadWebSite(chromeDriver)
    current_window = chromeDriver.current_window_handle
    for chrome_window in chromeDriver.window_handles:
        chromeDriver.switch_to.window(chrome_window)
        if chrome_window != current_window and chromeDriver.title == 'Options - Proxy Auto Auth':
            waitLoadWebSite(chromeDriver)
            sleep(3)
            findElementByInfo(chromeDriver, 'xpath=//input[@id="login"]').clear()
            userName = findElementByInfo(chromeDriver, 'xpath=//input[@id="login"]')
            setValueInput(userName, infoProxy[2])
            sleep(2)
            findElementByInfo(chromeDriver, 'xpath=//input[@id="password"]').clear()
            password = findElementByInfo(chromeDriver, 'xpath=//input[@id="password"]')
            sleep(2)
            setValueInput(password, infoProxy[3])
            sleep(2)
            btnSave = findElementByInfo(chromeDriver, 'xpath=//button[text()="Save"]')
            clickButton(btnSave)
            Sleep(2)
            chromeDriver.close()
            sleep(2)
            chromeDriver.switch_to.window(current_window)
            waitLoadWebSite(chromeDriver)
            chromeDriver.refresh()
            sleep(2)
            break

def waitLoadWebSite(chromeDriver):
    while chromeDriver.execute_script('return document.readyState') != 'complete':
        Sleep(2)

def waitLoadFinish(chromeDriver):
    while chromeDriver.execute_script('return document.readyState') != 'complete':
        Sleep(1)
    Sleep(0.1)

def goToWebsite(chromeDriver, link):
    chromeDriver.get(link)
    waitLoadFinish(chromeDriver)

def getListExtension(exPath: str):
    listSub = os.listdir(exPath)
    listPath = ""
    total = len(listSub)
    for index, folder in enumerate(listSub):
        listPath = listPath + exPath + "/" + folder
        if index < total - 1:
            listPath = listPath + ","
    return listPath

def switchToPage(chromeDriver, posPage):
    allHandles = chromeDriver.window_handles
    if posPage < len(allHandles):
        chromeDriver.switch_to.window(allHandles[posPage])
    sleep(1)

def waitOpenTab(chromeDriver):
    try:
        WebDriverWait(chromeDriver, TimeOut).until(EC.number_of_windows_to_be(2))
    except Exception as ex:
        Utils.writeError(str(ex))

def closePage(chromeDriver, posClose, posGo):
    allHandles = chromeDriver.window_handles
    if posClose < len(allHandles):
        chromeDriver.switch_to.window(allHandles[posClose])
        chromeDriver.close()
    if posGo < len(allHandles):
        chromeDriver.switch_to.window(allHandles[posGo])
    sleep(1)

def moveToElement(chromeDriver, elem):
    try:
        if elem != "":
            actions = ActionChains(chromeDriver)
            actions.move_to_element(elem).perform()
        else:
            Utils.writeError("moveToElement: Element null")
    except Exception as ex:
        Utils.writeError(str(ex))

def moveToElementandClick(chromeDriver, elem):
    try:
        if elem != "":
            actions = ActionChains(chromeDriver)
            actions.move_to_element(elem).perform()
            clickButton(elem)
        else:
            Utils.writeError("moveToElement: Element null")
    except Exception as ex:
        Utils.writeError(str(ex))

def dragAndDrop(chromeDriver, elem1, elem2):
    action_chains = ActionChains(chromeDriver)
    action_chains.drag_and_drop(elem1, elem2).perform()

def openWindow(chromeDriver, openURL):
    chromeDriver.execute_script("window.open('" + openURL + "', 'new_window')")

def openNewTab(chromeDriver, urlOpen):
    body = findElementByTagName(chromeDriver, "body")
    if body != "":
        body.send_keys(Keys.CONTROL + 't')
        body.send_keys(Keys.CONTROL + Keys.TAB)
        chromeDriver.get(urlOpen)

def scrollTo(chromeDriver, posXY):
    chromeDriver.execute_script("window.scrollBy(0," + str(posXY) + ");")

def scrollBy(chromeDriver, x, y):
    chromeDriver.execute_script("window.scrollBy("+str(x)+", "+str(y)+");")

def executeScript(chromeDriver, scriptRun):
    chromeDriver.execute_script(scriptRun)

def setValueInput(elem, dataInput):
    try:
        if elem != "":
            elem.send_keys(dataInput)
        else:
            Utils.writeError("setValueInput: Element null")
    except Exception as ex:
        Utils.writeError(ex.__class__)

def clickButton(elem):
    if elem != "":
        elem.click()
    else:
        Utils.writeError("clickButton: Element null")

def findElementByTagName(chromeDriver, tagName):
    element = ""
    # 0.5s
    try:
        element = WebDriverWait(chromeDriver, TimeOut).until(
            EC.presence_of_element_located((By.TAG_NAME, tagName))
        )
    except NoSuchElementException as noE:
        Utils.writeError(noE.__class__)
    except TimeoutException as tE:
        Utils.writeError(tE.__class__)
    except Exception as ex:
        Utils.writeError(ex.__class__)
    return element

def findElementsByTagName(chromeDriver, tagName):
    element = ""
    # 0.5s
    try:
        element = WebDriverWait(chromeDriver, TimeOut).until(
            EC.presence_of_all_elements_located((By.TAG_NAME, tagName))
        )
    except NoSuchElementException as noE:
        Utils.writeError(noE.__class__)
    except TimeoutException as tE:
        Utils.writeError(tE.__class__)
    except Exception as ex:
        Utils.writeError(ex.__class__)
    return element

def findElementById(chromeDriver, typeName, delayTime = TimeOut):
    element = ""
    try:
        element = WebDriverWait(chromeDriver, delayTime).until(
            EC.presence_of_element_located((By.ID, typeName))
        )
    except NoSuchElementException as noE:
        Utils.writeError(noE.__class__)
    except TimeoutException as tE:
        Utils.writeError(tE.__class__)
    except Exception as ex:
        Utils.writeError(ex.__class__)
    return element

def findElementByClass(chromeDriver, typeName, timeDelay = TimeOut):
    element = ""
    try:
        element = WebDriverWait(chromeDriver, timeDelay).until(
            EC.presence_of_element_located((By.CLASS_NAME, typeName))
        )
    except NoSuchElementException as noE:
        Utils.writeError(noE.__class__)
    except TimeoutException as tE:
        Utils.writeError(tE.__class__)
    except Exception as ex:
        Utils.writeError(ex.__class__)
    return element

def findElementByName(chromeDriver, typeName):
    element = ""
    try:
        element = WebDriverWait(chromeDriver, TimeOut).until(
            EC.presence_of_element_located((By.TAG_NAME, typeName))
        )
    except NoSuchElementException as noE:
        Utils.writeError(noE.__class__)
    except TimeoutException as tE:
        Utils.writeError(tE.__class__)
    except Exception as ex:
        Utils.writeError(ex.__class__)
    return element

def findElementsByXpath(chromeDriver, typeName, delayTime=TimeOut):
    elements = ""
    try:
        elements = WebDriverWait(chromeDriver, delayTime).until(
            EC.presence_of_all_elements_located((By.XPATH, typeName))
        )
    except NoSuchElementException as noE:
        Utils.writeError(noE.__class__)
    except TimeoutException as tE:
        Utils.writeError(tE.__class__)
    except Exception as ex:
        Utils.writeError(ex.__class__)
    return elements

def findElementByXpath(chromeDriver, typeName, delayTime=TimeOut):
    element = ''
    try:
        element = WebDriverWait(chromeDriver, delayTime).until(
            EC.presence_of_element_located((By.XPATH, typeName))
        )
    except NoSuchElementException as noE:
        Utils.writeError(noE.__class__)
    except TimeoutException as tE:
        Utils.writeError(tE.__class__)
    except Exception as ex:
        Utils.writeError(ex.__class__)
    return element

def findElementByLink(chromeDriver, linkName):
    element = ""
    try:
        element = WebDriverWait(chromeDriver, TimeOut).until(
            EC.presence_of_element_located((By.LINK_TEXT, linkName))
        )
    except NoSuchElementException as noE:
        Utils.writeError(noE.__class__)
    except TimeoutException as tE:
        Utils.writeError(tE.__class__)
    except Exception as ex:
        Utils.writeError(ex.__class__)
    return element

def findElementByCssSelector(chromeDriver, typeName):
    element = ""
    try:
        element = WebDriverWait(chromeDriver, TimeOut).until(
            EC.presence_of_element_located((By.CSS_SELECTOR, typeName))
        )
    except NoSuchElementException as noE:
        Utils.writeError(noE.__class__)
    except TimeoutException as tE:
        Utils.writeError(tE.__class__)
    except Exception as ex:
        Utils.writeError(ex.__class__)
    return element

def findElementByInfo(chromeDriver, target):
    element = ""
    if target.find("xpath") >= 0:
        xpath = target.replace("xpath=", "")
        element = findElementByXpath(chromeDriver, xpath)
    elif target.find("link") >= 0:
        link = target.replace("link=", "")
        element = findElementByLink(chromeDriver, link)
    if element != "":
        actions = ActionChains(chromeDriver)
        actions.move_to_element(element).perform()
    return element

def findElementByInfo(chromeDriver, target):
    element = ""
    if target.find("xpath") >= 0:
        xpath = target.replace("xpath=", "")
        element = findElementByXpath(chromeDriver, xpath)
    elif target.find("link") >= 0:
        link = target.replace("link=", "")
        element = findElementByLink(chromeDriver, link)
    if element != "":
        actions = ActionChains(chromeDriver)
        actions.move_to_element(element).perform()
    return element

def Sleep(timeSleep):
    sleep(timeSleep + random.randint(10, 200) / 1000)
